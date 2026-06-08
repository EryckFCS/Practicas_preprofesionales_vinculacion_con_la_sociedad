import os
import openpyxl

target_file = "/home/erick-fcs/Descargas/Practicas_preprofesionales/RESULTADOS-24-46 ERICK/25. IND-Derecho 2023.xlsx"

wb = openpyxl.load_workbook(target_file, data_only=True)
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    for r_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
        row_str = " | ".join(str(val) for val in row if val is not None)
        if any(keyword in row_str.lower() for keyword in ["total", "población", "poblacion", "graduados", "encuestados"]):
            print(f"[{sheet_name} Row {r_idx}]: {row_str}")
