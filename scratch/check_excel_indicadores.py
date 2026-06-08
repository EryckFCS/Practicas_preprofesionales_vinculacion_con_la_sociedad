import os
import openpyxl
from src.config import CAREER_FILES, RESULTADOS_DIR

for rid in sorted(CAREER_FILES.keys()):
    if rid >= 35:
        path = os.path.join(RESULTADOS_DIR, CAREER_FILES[rid]["excel"])
        wb = openpyxl.load_workbook(path, data_only=True)
        # Find sheet name starting with 'Indicadores'
        sheet_name = [s for s in wb.sheetnames if s.strip().startswith('Indicadores')][0]
        sheet = wb[sheet_name]
        # Look for Total row or check row 31 (which is index 31)
        row_31 = [sheet.cell(31, c).value for c in range(1, 10)]
        print(f"Report {rid}: sheet='{sheet_name}', row_31={row_31}")
