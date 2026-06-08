import os
import openpyxl

# Find the file
target_file = None
for root, dirs, files in os.walk("/home/erick-fcs/Descargas/Practicas_preprofesionales"):
    for file in files:
        if file.startswith(".~lock."):
            continue
        if "25. IND-Derecho 2023.xlsx" in file or "25. IND-Derecho" in file:
            target_file = os.path.join(root, file)
            break
    if target_file:
        break

print("Found file:", target_file)

if target_file:
    wb = openpyxl.load_workbook(target_file, data_only=True)
    print("Sheets:", wb.sheetnames)
    
    # Search for 89 and 68
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for r_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
            for c_idx, val in enumerate(row, 1):
                if val == 89 or val == "89":
                    print(f"Found 89 in sheet '{sheet_name}' at row {r_idx}, col {c_idx}")
                elif val == 68 or val == "68":
                    print(f"Found 68 in sheet '{sheet_name}' at row {r_idx}, col {c_idx}")
